"""
AGEN-OPS-6 — SME-AI workflow benchmark (v0.5).

Tests open-source and closed-source LLMs on eight workflow variants covering
real-world SME operations tasks. Two backends: Tensorix (open models) and
OpenRouter (VLMs + closed-model anchors like GPT-5.5 and Claude Opus).

Default workflows (--workflow all):
    compliance      Material certificate compliance (text)
    dims            Drawing-vs-3D dimension comparison reasoning (text)
    dims_vlm        Direct VLM dimension extraction from engineering drawings
    docs            PDF entity extraction (native + OCR fallback)
    schedule_read   Spreadsheet issue detection
    schedule_write  Spreadsheet cascade-update editing

Additional (run explicitly via --workflow):
    dims_ocr        EasyOCR + LLM hybrid dim extraction (appendix)
    docs_vlm        Direct VLM PDF extraction (appendix)

Each (workflow, model[, drawing/PDF]) combination is run N times (default 10)
and aggregated with mean / std / CI / min / max for precision, recall, F1, latency.
Infrastructure failures (timeouts, 402, 404) are separated from capability failures.
Transient errors (429, 502, 503) are retried automatically with exponential backoff.

Setup:
    uv sync
    cp .env.example .env   # fill in API keys (see .env.example)

Run:
    uv run python run_experiment.py                              # all default workflows, 10 runs
    uv run python run_experiment.py --workflow compliance         # one workflow
    uv run python run_experiment.py --runs 1 --check-credits     # smoke test with credit check
    uv run python run_experiment.py --workflow dims_ocr           # appendix workflow

Outputs:
    results/run_<timestamp>.json        full raw runs + aggregates
    results/summary_<timestamp>.md      side-by-side overview table
    stdout                              pretty-printed summary
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import signal
import statistics
import sys
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).parent
PROMPTS = ROOT / "prompts"
DATA = ROOT / "data"
RESULTS = ROOT / "results"
OCR_CACHE = DATA / "dims" / "ocr_cache"
RESULTS.mkdir(exist_ok=True)
OCR_CACHE.mkdir(exist_ok=True)

# Silence PyTorch / EasyOCR boot noise. Forcing CUDA_VISIBLE_DEVICES="" makes
# torch skip CUDA initialisation entirely (no driver-version warning), and the
# warnings filter swallows the residual UserWarnings.
import warnings as _warnings
os.environ.setdefault("CUDA_VISIBLE_DEVICES", "")
os.environ.setdefault("KMP_DUPLICATE_LIB_OK", "TRUE")
os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")
_warnings.filterwarnings("ignore", category=UserWarning)
_warnings.filterwarnings("ignore", category=FutureWarning)


# ----------------------------- env loader --------------------------------- #

def load_env_file(path: Path) -> None:
    """Minimal .env loader: KEY=VALUE per line, # comments, blank lines ignored.
    Existing environment variables are NOT overridden (so a real env var still wins)."""
    if not path.is_file():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip()
        if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
            value = value[1:-1]
        if key and key not in os.environ:
            os.environ[key] = value


load_env_file(ROOT / ".env")

try:
    from openai import OpenAI
except ImportError:
    print("ERROR: openai package not installed. Run: uv sync", file=sys.stderr)
    sys.exit(2)

import random as _random


# Per-API-call timeout. We saw individual gpt-oss-20b runs spend 188s and 370s on
# docs because the OpenAI SDK's default max_retries=2 turned a single timeout
# into a chain of retries. RUN_TIMEOUT_S caps each attempt; max_retries=0 on the
# client (see ClientPool) prevents the SDK from auto-retrying; and call_model /
# call_vlm skip the plain-text fallback if the json_object attempt hit a timeout.
# Net effect: each run is bounded at ~RUN_TIMEOUT_S wall clock.
RUN_TIMEOUT_S = 150.0


def _is_timeout_error(e: Exception) -> bool:
    """Heuristic for 'this exception was a timeout'. We use it to decide whether to
    fall back to the plain-text retry path: if the json_object attempt timed out,
    a second attempt would just burn another RUN_TIMEOUT_S seconds. Skip it."""
    s = (str(e) or "").lower()
    cls = e.__class__.__name__.lower()
    return (
        ("timeout" in s)
        or ("timed out" in s)
        or ("apitimeouterror" in cls)
        or ("hardtimeout" in cls)
    )


class HardTimeoutError(Exception):
    """Raised when our SIGALRM-based hard wall-clock timeout fires. We use this as
    a backstop because the OpenAI SDK's `timeout=` parameter has been observed to
    hang indefinitely on half-open connections / slow OpenRouter providers."""


def _alarm_handler(signum, frame):
    raise HardTimeoutError(f"hard wall-clock timeout fired after {RUN_TIMEOUT_S * 1.2:.0f}s")


@contextmanager
def hard_timeout(seconds: float):
    """SIGALRM-based hard kill that bypasses the network stack. Unix-only.
    On Windows / non-main threads SIGALRM is not available; we silently skip
    the watchdog there (the SDK timeout is the only line of defence in that case).
    """
    use_alarm = (
        hasattr(signal, "SIGALRM")
        and threading_main_thread()
    )
    old = None
    if use_alarm:
        old = signal.signal(signal.SIGALRM, _alarm_handler)
        # signal.alarm only accepts integer seconds. Round up.
        signal.alarm(max(1, int(seconds) + 1))
    try:
        yield
    finally:
        if use_alarm:
            signal.alarm(0)
            if old is not None:
                signal.signal(signal.SIGALRM, old)


def threading_main_thread() -> bool:
    """Return True if we're running on the main thread (signal.alarm needs this)."""
    return threading.current_thread() is threading.main_thread()


# ----------------------------- transient-error retry ----------------------- #

# Transient HTTP errors (429 rate-limit, 502/503 bad gateway, Cloudflare 520/522/524)
# are worth retrying with backoff. Fatal errors (402 payment, 404 not found, 401/403
# auth) should fail immediately. Timeouts are handled separately by the existing
# circuit-breaker logic and are NOT retried here.
MAX_TRANSIENT_RETRIES = 2
TRANSIENT_BACKOFF_BASE = 2.0  # seconds; actual wait = base * 2^attempt → 2s, 4s


def _is_transient_error(e: Exception) -> bool:
    """Return True for HTTP errors that are worth retrying with backoff.

    Covers rate limits (429), bad gateway (502/503), Cloudflare errors (520/522/524),
    and connection-level failures (reset, refused, aborted)."""
    status = getattr(e, "status_code", None)
    if status in (429, 502, 503, 520, 522, 524):
        return True
    s = str(e).lower()
    cls = e.__class__.__name__.lower()
    if "apiconnectionerror" in cls or "connectionerror" in cls:
        return True
    if "connection" in s and ("reset" in s or "refused" in s or "aborted" in s):
        return True
    return False


def _is_fatal_api_error(e: Exception) -> bool:
    """Return True for API errors that should NOT be retried (auth, payment, not-found)."""
    status = getattr(e, "status_code", None)
    return status in (401, 402, 403, 404)


def _model_extra_kwargs(model: str) -> dict:
    """Return per-model kwargs to merge into chat.completions.create().

    Handles:

    * gpt-oss-*  uses the harmony format with a separate reasoning channel.
                 Set reasoning_effort="low" to skip extended reasoning.
    * qwen3*/qwen3.5* support an `enable_thinking` chat-template flag.
                 Disable it to avoid burning output budget on internal reasoning.
    * OpenRouter routes the same model to multiple backend providers, some of
                 which serve heavily-quantized fast copies that return empty
                 results in 1-5 seconds while quality providers take 60-100s.
                 Constrain routing with relaxed quantization filters (int8 allowed
                 after first-run 404s) and allow fallbacks.
    * Closed-source anchors (GPT-5.5, Claude Opus) routed through OpenRouter
                 with minimal constraints (no quantization filter — they're served
                 by the original provider, not a quantized replica).

    Also injects a per-call seed to defeat any request-level caching the
    provider might do.
    """
    backend = MODEL_BACKEND.get(model, "tensorix")

    extras: dict = {}
    extras["seed"] = int(time.time() * 1_000_000) & 0x7FFFFFFF
    extra_body: dict = {}

    # OpenRouter provider-routing — relaxed from v0.4 (added int8 to fix 404s
    # on VLM endpoints; removed require_parameters for VLM models where it
    # caused routing failures because backends don't advertise seed support).
    # Closed-source anchors (openai/gpt-5*, anthropic/*) skip quantization
    # filters entirely — they're served by the original provider.
    if backend == "openrouter":
        is_closed = model.startswith("openai/gpt-5") or model.startswith("anthropic/")
        is_vlm = any(tag in model for tag in ("-vl-", "-vl ", "vlm", "-vision", "vision-", "qwen3.5-122b", "qwen3.5-397b"))
        if is_closed:
            extra_body["provider"] = {
                "require_parameters": False,
                "allow_fallbacks": True,
            }
        else:
            extra_body["provider"] = {
                "require_parameters": not is_vlm,   # relax for VLMs (seed not always advertised)
                "quantizations": ["fp16", "bf16", "fp8", "int8"],  # allow int8 (fixes first-run 404s)
                "allow_fallbacks": True,
            }

    if model.startswith("openai/gpt-oss"):
        extra_body["reasoning_effort"] = "low"
    elif model.startswith("qwen/qwen3") or model.startswith("qwen/qwen-3"):
        extra_body["chat_template_kwargs"] = {"enable_thinking": False}
        extra_body["enable_thinking"] = False

    if extra_body:
        extras["extra_body"] = extra_body
    return extras


# Per-model backend routing. "tensorix" uses TENSORIX_API_KEY/BASE_URL,
# "openrouter" uses OPENROUTER_API_KEY/BASE_URL. Both are OpenAI-compatible.
# (qwen/qwen-2.5-coder-32b-instruct removed: as a coder-tuned 32B it consistently
# emitted malformed JSON on the structured-output prompts. Tier-0.5 comparator
# is now represented by openai/gpt-oss-20b.)
MODEL_BACKEND: dict[str, str] = {
    # Tensorix-hosted
    "meta-llama/llama-3.3-70b-instruct": "tensorix",
    "qwen/qwen-2.5-72b-instruct":        "tensorix",
    "z-ai/glm-5.1":                      "tensorix",
    "openai/gpt-oss-120b":               "tensorix",
    "openai/gpt-oss-20b":                "tensorix",
    "qwen/qwen3.5-9b":                   "tensorix",
    "deepseek/deepseek-v4-flash":        "tensorix",
    "meta-llama/llama-4-maverick":       "tensorix",
    "qwen/qwen3-235b-a22b-2507":         "tensorix",
    # All vision-language models routed through OpenRouter. Tensorix's hosting of
    # qwen3-vl-235b and qwen3.5-122b was either silently caching (breaking variance
    # estimation) or timing out under the 150s budget — OpenRouter's hosting of the
    # same models is more reliable for our purposes.
    "qwen/qwen3-vl-235b-a22b-instruct":  "openrouter",  # Qwen3-VL 235B MoE, ~22B active
    "qwen/qwen3.5-122b-a10b":            "openrouter",  # Qwen 3.5 multimodal-native, ~10B active
    "qwen/qwen-2.5-vl-72b-instruct":     "openrouter",  # baseline (last run F1≈0.77)
    "qwen/qwen3.5-397b-a17b":            "openrouter",  # flagship multimodal-native, ~17B active
    "qwen/qwen3-vl-235b-a22b-thinking":  "openrouter",  # 235B Thinking variant — STEM/technical
    "qwen/qwen3-vl-30b-a3b-instruct":    "openrouter",  # 30B MoE VL, ~3B active — Tier 0.5
    # Anchor models — closed-source reference ceilings, routed via OpenRouter.
    "openai/gpt-5.5":                   "openrouter",     # GPT-5.5 via OpenRouter
    "anthropic/claude-opus-4.6":        "openrouter",     # Claude Opus 4.6 via OpenRouter
}

# Workflow-aware defaults (applied when --models is NOT passed).
# Text workflows compare a full ladder from 9B → 120B so we can read off the
# minimum hardware tier where the workflow still works. The extract workflow
# keeps Qwen 2.5 72B as the primary baseline plus three downside comparators.
# Anchor (closed-source) models — reference ceilings, not local-deployment candidates.
_ANCHOR_MODELS = [
    "openai/gpt-5.5",                      # GPT-5.5 via OpenRouter
    "anthropic/claude-opus-4.6",           # Claude Opus 4.6 via OpenRouter
]

_FULL_LADDER = [
    *_ANCHOR_MODELS,                       # closed-model reference ceilings
    "qwen/qwen3-235b-a22b-2507",           # ~235B MoE, ~22B active — Tier 2+ (July 2025)
    "openai/gpt-oss-120b",                 # ~120B MoE — Tier 2 only
    "meta-llama/llama-4-maverick",         # ~400B MoE, ~17B active — Tier 2
    "qwen/qwen-2.5-72b-instruct",          # 72B — Tier 1+
    "meta-llama/llama-3.3-70b-instruct",   # 70B — Tier 1+
    "deepseek/deepseek-v4-flash",          # MoE — size unverified
    "z-ai/glm-5.1",                        # capable open-weights (size unverified)
    "openai/gpt-oss-20b",                  # 20B — Tier 0.5 representative
    "qwen/qwen3.5-9b",                     # 9B — Tier 0
]
_VISION_LADDER = [
    "qwen/qwen-2.5-72b-instruct",          # primary baseline
    "openai/gpt-oss-120b",                 # upper bound
    "openai/gpt-oss-20b",                  # mid downside
    "qwen/qwen3.5-9b",                     # low downside
]

# Direct vision-language ladder for dims_vlm and docs_vlm (image / page images go
# straight to a VLM, no OCR step). All routed via OpenRouter for consistent
# behaviour — Tensorix's VLM hosting was silently caching and/or timing out.
# Hardware tier annotations refer to what the model would need if SELF-HOSTED;
# OpenRouter is just the test harness for now.
# Anchor models (GPT-5.5, Opus 4.6) are included — both support vision natively.
_VLM_LADDER = [
    *_ANCHOR_MODELS,                        # closed-model reference ceilings (both support vision)
    "qwen/qwen3-vl-235b-a22b-instruct",    # ~140 GB Q4 — Tier 2.5 (235B MoE, ~22B active)
    "qwen/qwen3.5-122b-a10b",              # ~70 GB Q4  — Tier 2   (multimodal-native, ~10B active)
    "qwen/qwen3.5-397b-a17b",              # ~225 GB Q4 — Tier 3   (flagship, ~17B active)
    "qwen/qwen3-vl-235b-a22b-thinking",    # ~140 GB Q4 — Tier 2.5 (Thinking variant)
    "qwen/qwen-2.5-vl-72b-instruct",       # ~40 GB Q4  — Tier 1+  (baseline; prior F1≈0.77)
    "qwen/qwen3-vl-30b-a3b-instruct",      # ~18 GB Q4  — Tier 0.5 (30B MoE VL, ~3B active)
]

DEFAULT_MODELS_PER_WORKFLOW: dict[str, list[str]] = {
    "compliance":     list(_FULL_LADDER),
    "dims":           list(_FULL_LADDER),
    "dims_ocr":       list(_VISION_LADDER),
    "dims_vlm":       list(_VLM_LADDER),
    "docs":           list(_FULL_LADDER),
    "docs_vlm":       list(_VLM_LADDER),
    "schedule_read":  list(_FULL_LADDER),
    "schedule_write": list(_FULL_LADDER),
}

# Flat union (used when user passes --models, applied to every workflow).
DEFAULT_MODELS = sorted({m for ms in DEFAULT_MODELS_PER_WORKFLOW.values() for m in ms})


class ClientPool:
    """Lazy multi-backend client factory.

    Tensorix and OpenRouter both speak the OpenAI Chat Completions schema, so we
    keep one OpenAI() instance per backend. Anchor models (GPT-5.5, Claude Opus)
    are routed through OpenRouter — no separate API keys needed.
    """

    def __init__(self) -> None:
        self._tensorix = None
        self._openrouter = None
        self._lock = threading.Lock()

    def _backend_for(self, model: str) -> str:
        return MODEL_BACKEND.get(model, "tensorix")

    def get(self, model: str):
        backend = self._backend_for(model)
        # max_retries=0 disables the OpenAI SDK's automatic retry-on-timeout behaviour
        # (default = 2). Without this, a single 150s timeout becomes ~450s wall clock.
        # Our own _create_with_retry handles transient-error retries selectively.
        with self._lock:
            if backend == "openrouter":
                if self._openrouter is None:
                    key = os.environ.get("OPENROUTER_API_KEY")
                    url = os.environ.get("OPENROUTER_BASE_URL", "https://openrouter.ai/api/v1")
                    if not key:
                        raise RuntimeError(f"OPENROUTER_API_KEY not set (needed for {model}). Add it to .env.")
                    self._openrouter = OpenAI(api_key=key, base_url=url, max_retries=0)
                return self._openrouter
            # default: tensorix
            if self._tensorix is None:
                key = os.environ.get("TENSORIX_API_KEY")
                url = os.environ.get("TENSORIX_BASE_URL")
                if not key or not url:
                    raise RuntimeError("TENSORIX_API_KEY or TENSORIX_BASE_URL not set. See .env.example.")
                self._tensorix = OpenAI(api_key=key, base_url=url, max_retries=0)
            return self._tensorix


# ----------------------------- client pool --------------------------------- #

def make_client_pool() -> ClientPool:
    """Create the two-backend pool (Tensorix + OpenRouter). Validates Tensorix
    env vars eagerly because most open-model runs go through Tensorix; OpenRouter
    is validated lazily on first use. Anchor models (GPT-5.5, Claude Opus) and all
    VLM models are routed through OpenRouter."""
    if not os.environ.get("TENSORIX_API_KEY") or not os.environ.get("TENSORIX_BASE_URL"):
        print("WARNING: TENSORIX_API_KEY or TENSORIX_BASE_URL not set.", file=sys.stderr)
        print("         Tensorix-hosted models will fail. See .env.example.", file=sys.stderr)
        print()
    return ClientPool()


_INTERESTING_HEADERS = [
    "x-provider", "openrouter-processing-ms", "x-request-id",
    "x-ratelimit-limit-requests", "x-ratelimit-remaining-requests",
    "x-ratelimit-limit-tokens", "x-ratelimit-remaining-tokens",
]


def _create_with_headers(client, create_kwargs: dict) -> tuple:
    """Call chat.completions.create, capturing response headers when possible.

    Returns (response, headers_dict).
    """
    headers: dict[str, str] = {}
    raw = client.chat.completions.with_raw_response.create(**create_kwargs)
    for h in _INTERESTING_HEADERS:
        v = raw.headers.get(h)
        if v:
            headers[h] = v
    resp = raw.parse()
    if hasattr(resp, 'model') and resp.model:
        headers["served_model"] = resp.model
    return resp, headers


def _create_with_retry(client, create_kwargs: dict) -> tuple:
    """Wrap _create_with_headers with retries for transient HTTP errors.

    Retries up to MAX_TRANSIENT_RETRIES times on 429/502/503/connection errors
    with exponential backoff (2s, 4s). Fatal errors (402, 404) and timeouts
    propagate immediately without consuming retry budget.

    Returns (response, headers, retry_log) where retry_log is a list of
    transient failures that were retried before the final attempt.

    Wall-clock impact: transient errors return quickly (seconds, not minutes),
    so retries typically add <10s total. The per-call SDK timeout still bounds
    each individual attempt.
    """
    retry_log: list[dict] = []
    last_err: Exception | None = None
    for attempt in range(1 + MAX_TRANSIENT_RETRIES):
        try:
            resp, headers = _create_with_headers(client, create_kwargs)
            return resp, headers, retry_log
        except Exception as e:
            last_err = e
            if _is_timeout_error(e) or _is_fatal_api_error(e):
                raise
            if _is_transient_error(e) and attempt < MAX_TRANSIENT_RETRIES:
                wait = TRANSIENT_BACKOFF_BASE * (2 ** attempt)
                retry_log.append({
                    "attempt": attempt + 1,
                    "error": str(e)[:200],
                    "wait_s": round(wait, 1),
                })
                time.sleep(wait)
                continue
            raise
    raise last_err  # unreachable, but keeps type checkers happy


def call_model(client, model: str, system_prompt: str, user_prompt: str, temperature: float) -> tuple[str, dict]:
    """Call the model. Try response_format=json_object first; fall back to plain text.
    On a timeout we skip the fallback — see RUN_TIMEOUT_S note above. Net wall clock
    per call is bounded at ~RUN_TIMEOUT_S (timeout fires fast) or ~2 × RUN_TIMEOUT_S
    (json_object refused for non-timeout reasons, then plain-text retry).

    Transient HTTP errors (429, 502, 503) are retried up to MAX_TRANSIENT_RETRIES
    times with exponential backoff before propagating.

    Captures provider response headers (x-provider, processing time, rate limits)
    when available, stored in metadata["response_headers"].
    """
    metadata: dict = {"attempts": []}
    last_was_timeout = False
    extra_kwargs = _model_extra_kwargs(model)

    create_kwargs = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "response_format": {"type": "json_object"},
        "timeout": RUN_TIMEOUT_S,
        **extra_kwargs,
    }

    try:
        t0 = time.time()
        with hard_timeout(RUN_TIMEOUT_S * 1.2):
            resp, headers, retries = _create_with_retry(client, create_kwargs)
        elapsed = time.time() - t0
        if headers:
            metadata["response_headers"] = headers
        if retries:
            metadata["transient_retries"] = retries
        content = resp.choices[0].message.content
        if content is not None and content.strip():
            metadata["attempts"].append({"mode": "json_object", "elapsed_s": round(elapsed, 2), "ok": True})
            return content, metadata
        metadata["attempts"].append({"mode": "json_object", "elapsed_s": round(elapsed, 2), "ok": False, "error": "empty content under json_object mode"})
    except Exception as e:
        last_was_timeout = _is_timeout_error(e)
        metadata["attempts"].append({"mode": "json_object", "ok": False, "error": str(e)[:200], "was_timeout": last_was_timeout})
        if last_was_timeout:
            raise

    create_kwargs_plain = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "timeout": RUN_TIMEOUT_S,
        **extra_kwargs,
    }

    t0 = time.time()
    with hard_timeout(RUN_TIMEOUT_S * 1.2):
        resp, headers, retries = _create_with_retry(client, create_kwargs_plain)
    elapsed = time.time() - t0
    if headers:
        metadata["response_headers"] = headers
    if retries:
        metadata.setdefault("transient_retries", []).extend(retries)
    metadata["attempts"].append({"mode": "plain", "elapsed_s": round(elapsed, 2), "ok": True})
    return resp.choices[0].message.content, metadata


def parse_json_response(text: str) -> dict:
    """Extract JSON from the model output, tolerant of code-fenced or wrapped responses,
    None responses (which happen on some providers when the model refuses or hits a
    safety filter), and empty strings."""
    if text is None:
        raise ValueError("model returned None — likely a provider-side refusal, safety filter, or empty response")
    text = text.strip()
    if not text:
        raise ValueError("model returned empty response")
    if text.startswith("```"):
        lines = text.split("\n")
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        text = "\n".join(lines).strip()
    return json.loads(text)


# ----------------------------- compliance evaluator ----------------------- #

def run_compliance(client: OpenAI, model: str, temperature: float) -> dict:
    system_prompt = (PROMPTS / "compliance_system.txt").read_text(encoding="utf-8")
    spec = (DATA / "compliance" / "customer_spec.txt").read_text(encoding="utf-8")
    cert = (DATA / "compliance" / "supplier_cert.txt").read_text(encoding="utf-8")
    ground_truth = json.loads((DATA / "compliance" / "ground_truth.json").read_text(encoding="utf-8"))

    user_prompt = (
        "CUSTOMER SPECIFICATION:\n===================================================\n"
        f"{spec}\n\nSUPPLIER CERTIFICATE:\n===================================================\n"
        f"{cert}\n\nCross-check the certificate against the specification and produce the JSON object as instructed."
    )

    raw, meta = call_model(client, model, system_prompt, user_prompt, temperature)

    try:
        parsed = parse_json_response(raw)
    except Exception as e:
        return {"workflow": "compliance", "model": model, "raw_output": raw, "error": f"JSON parse failed: {e}", "metadata": meta}

    flagged = parsed.get("deviations", []) if isinstance(parsed, dict) else []

    expected_devs = ground_truth["expected_deviations"]
    expected_keywords = {
        "DEV-1": ["yield strength"],
        # Avoid bare "ra" (substring of "intergranular" etc.); use distinctive phrases.
        "DEV-2": ["surface roughness", "surface finish", "ra ", " ra,", " ra(", "ra value", "0.95", "0.8 micrometre", "0.8 µm"],
        "DEV-3": ["intergranular", "a262"],
        "DEV-4": ["charpy", "impact at -196", "impact toughness"],
        "DEV-5": ["a276-17", "a276-23", "a479-23", "superseded", "revision", "standard version", "outdated standard"],
    }

    flagged_normalised = [
        ((f.get("field", "") or "") + " " + (f.get("rationale", "") or "") + " " + (f.get("spec_requirement", "") or "")).lower()
        for f in flagged
    ]

    caught: set[str] = set()
    for dev_id, keywords in expected_keywords.items():
        for entry in flagged_normalised:
            if any(k in entry for k in keywords):
                caught.add(dev_id)
                break

    expected_kw_flat = [k for ks in expected_keywords.values() for k in ks]
    fp_count = 0
    fp_fields: list[str] = []
    for f, normalised in zip(flagged, flagged_normalised):
        if not any(k in normalised for k in expected_kw_flat):
            fp_count += 1
            fp_fields.append(f.get("field", "<unnamed>") if isinstance(f, dict) else "<unnamed>")

    n_caught = len(caught)
    n_total = len(expected_devs)
    precision = n_caught / max(1, n_caught + fp_count)
    recall = n_caught / n_total if n_total else 0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    return {
        "workflow": "compliance",
        "model": model,
        "temperature": temperature,
        "metadata": meta,
        "raw_output": raw,
        "score": {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "caught": sorted(caught),
            "missed": sorted(set(d["id"] for d in expected_devs) - caught),
            "false_positive_count": fp_count,
            "false_positive_fields": fp_fields,
        },
    }


# ----------------------------- dims evaluator ------------------------------ #

def load_csv(path: Path) -> list[dict]:
    with path.open() as f:
        return list(csv.DictReader(f))


def run_dims(client: OpenAI, model: str, temperature: float) -> dict:
    system_prompt = (PROMPTS / "dims_system.txt").read_text(encoding="utf-8")
    drawing = load_csv(DATA / "dims" / "drawing_dims.csv")
    model_dims = load_csv(DATA / "dims" / "model_dims.csv")
    ground_truth = json.loads((DATA / "dims" / "ground_truth.json").read_text(encoding="utf-8"))

    drawing_table = "DRAWING DIMENSIONS (extracted from 2D drawing):\ndim_id | feature | nominal_value | tolerance_plus | tolerance_minus\n"
    for r in drawing:
        drawing_table += f"{r['dim_id']} | {r['feature']} | {r['nominal_value']} | {r['tolerance_plus']} | {r['tolerance_minus']}\n"

    model_table = "\n3D MODEL DIMENSIONS (extracted from 3D model):\ndim_id | feature | model_value\n"
    for r in model_dims:
        model_table += f"{r['dim_id']} | {r['feature']} | {r['model_value']}\n"

    user_prompt = drawing_table + model_table + "\n\nReview each dimension and produce the JSON object as instructed. Apply the 10% tolerance band rule strictly."

    raw, meta = call_model(client, model, system_prompt, user_prompt, temperature)

    try:
        parsed = parse_json_response(raw)
    except Exception as e:
        return {"workflow": "dims", "model": model, "raw_output": raw, "error": f"JSON parse failed: {e}", "metadata": meta}

    flagged = parsed.get("flagged_dimensions", []) if isinstance(parsed, dict) else []
    flagged_ids = sorted({f.get("dim_id", "") for f in flagged if isinstance(f, dict)})
    must_flag_ids = sorted({d["dim_id"] for d in ground_truth["must_flag"]})

    tp = sorted(set(flagged_ids) & set(must_flag_ids))
    fp = sorted(set(flagged_ids) - set(must_flag_ids))
    fn = sorted(set(must_flag_ids) - set(flagged_ids))

    precision = len(tp) / max(1, len(tp) + len(fp))
    recall = len(tp) / max(1, len(tp) + len(fn))
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    return {
        "workflow": "dims",
        "model": model,
        "temperature": temperature,
        "metadata": meta,
        "raw_output": raw,
        "score": {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "true_positives": tp,
            "false_positives": fp,
            "false_negatives": fn,
        },
    }


# ---------------------- dims_ocr evaluator -------------------------------- #

_OCR_READER = None


def get_ocr_reader():
    global _OCR_READER
    if _OCR_READER is None:
        try:
            import easyocr  # type: ignore
        except ImportError:
            print("ERROR: easyocr not installed. Run: uv sync", file=sys.stderr)
            sys.exit(2)
        # English + German covers our drawings (EN labels) and scanned legal PDFs (DE).
        # Adding more languages here triggers another one-time model download per language pair.
        _OCR_READER = easyocr.Reader(["en", "de"], gpu=False, verbose=False)
    return _OCR_READER


def ocr_drawing(drawing_path: Path) -> list[dict]:
    """Run EasyOCR on a drawing image and cache the result on disk.

    Auto-detects corrupted / empty caches (e.g. from a previously interrupted run)
    and regenerates them. A valid cache is a JSON list of >=1 dicts.
    """
    cache_path = OCR_CACHE / (drawing_path.stem + ".json")
    if cache_path.is_file():
        try:
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            if isinstance(cached, list) and len(cached) > 0 and all(isinstance(x, dict) for x in cached):
                return cached
            print(f"  ⚠ OCR cache for {drawing_path.name} is empty / invalid — regenerating.")
        except Exception as e:
            print(f"  ⚠ OCR cache for {drawing_path.name} unreadable ({e}) — regenerating.")
        try:
            cache_path.unlink()
        except Exception:
            pass

    reader = get_ocr_reader()
    raw = reader.readtext(str(drawing_path))  # list of [bbox(4 corners), text, conf]
    cleaned = []
    for bbox, text, conf in raw:
        # bbox is a list of 4 [x, y] points
        xs = [p[0] for p in bbox]
        ys = [p[1] for p in bbox]
        cleaned.append({
            "text": text,
            "bbox_x_min": int(min(xs)),
            "bbox_y_min": int(min(ys)),
            "bbox_x_max": int(max(xs)),
            "bbox_y_max": int(max(ys)),
            "confidence": round(float(conf), 3),
        })
    cache_path.write_text(json.dumps(cleaned, indent=2), encoding="utf-8")
    return cleaned


def _norm_kind(s: str) -> str:
    """Normalise a dimension `kind` field. Returns one of:
        diameter, length, radius, angle, chamfer, thread, perpendicularity,
        surface_finish, undercut_callout, or the raw lowercased value if unknown.
    Note: diameter and diameter_pattern (hole patterns) collapse to "diameter" — the
    underlying value semantics are the same; kind alone shouldn't make match/miss.
    """
    s = (s or "").lower().strip()
    aliases = {
        "diameter": "diameter",
        "ø": "diameter",
        "od": "diameter",
        "id": "diameter",
        "diameter_pattern": "diameter",   # collapse: same numeric meaning
        "hole_pattern": "diameter",
        "length": "length",
        "linear": "length",
        "distance": "length",
        "radius": "radius",
        "r": "radius",
        "fillet": "radius",
        "angle": "angle",
        "chamfer": "chamfer",
        "thread": "thread",
        "perpendicularity": "perpendicularity",
        "gd&t": "perpendicularity",
        "gdt": "perpendicularity",
        "flatness": "perpendicularity",
        "surface_finish": "surface_finish",
        "surface finish": "surface_finish",
        "ra": "surface_finish",
        "undercut_callout": "undercut_callout",
        "undercut": "undercut_callout",
        "position": "position",
        "true_position": "position",
        "true position": "position",
        "tp": "position",
    }
    return aliases.get(s, s)


def _safe_float(v: Any) -> float | None:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _match_dim(extracted: dict, gt: dict, value_eps: float, tol_eps: float) -> bool:
    if _norm_kind(extracted.get("kind", "")) != _norm_kind(gt.get("kind", "")):
        return False
    ev = _safe_float(extracted.get("value"))
    gv = _safe_float(gt.get("value"))
    if ev is None or gv is None:
        # string callouts (e.g. undercut)
        return str(extracted.get("value", "")).strip().lower() == str(gt.get("value", "")).strip().lower()
    if abs(ev - gv) > value_eps:
        return False
    # If GT specifies tolerance, check it (within tol_eps).
    for side in ("tolerance_plus", "tolerance_minus"):
        if side in gt:
            etol = _safe_float(extracted.get(side, 0)) or 0
            gtol = _safe_float(gt.get(side, 0)) or 0
            if abs(etol - gtol) > tol_eps:
                return False
    return True


def run_dims_ocr(client: OpenAI, model: str, temperature: float, drawing_name: str) -> dict:
    drawing_path = DATA / "dims" / drawing_name
    gt_path = DATA / "dims" / f"{drawing_path.stem}_ground_truth.json"
    if not drawing_path.is_file() or not gt_path.is_file():
        return {"workflow": "dims_ocr", "drawing": drawing_name, "model": model, "error": "drawing or GT file missing"}

    ground_truth = json.loads(gt_path.read_text(encoding="utf-8"))
    ocr_items = ocr_drawing(drawing_path)

    system_prompt = (PROMPTS / "dims_ocr_system.txt").read_text(encoding="utf-8")

    ocr_text_block = "OCR OUTPUT (text, position, confidence) from the drawing image:\n"
    ocr_text_block += "idx | text | x_min | y_min | x_max | y_max | conf\n"
    for i, it in enumerate(ocr_items):
        ocr_text_block += f"{i} | {it['text']} | {it['bbox_x_min']} | {it['bbox_y_min']} | {it['bbox_x_max']} | {it['bbox_y_max']} | {it['confidence']}\n"

    drawing_meta = (
        f"\nDrawing metadata (for context):\n"
        f"  title: {ground_truth.get('title', '?')}\n"
        f"  material: {ground_truth.get('material', '?')}\n"
        f"  units: {ground_truth.get('units', '?')}\n"
        f"  scale: {ground_truth.get('scale', '?')}\n"
        f"  general tolerance note: {ground_truth.get('general_tolerance_note', '?')}\n"
    )

    user_prompt = ocr_text_block + drawing_meta + "\n\nFrom the OCR output above, extract every dimension and tolerance present on the drawing and return the JSON object as instructed."

    raw, meta = call_model(client, model, system_prompt, user_prompt, temperature)

    try:
        parsed = parse_json_response(raw)
    except Exception as e:
        return {"workflow": "dims_ocr", "drawing": drawing_name, "model": model, "raw_output": raw, "error": f"JSON parse failed: {e}", "metadata": meta}

    extracted = parsed.get("extracted_dimensions", []) if isinstance(parsed, dict) else []
    gt_dims = ground_truth.get("expected_dimensions", [])

    units = (ground_truth.get("units") or "").lower()
    value_eps = 0.05 if "milli" in units or "mm" in units else 0.01
    tol_eps = 0.05 if "milli" in units or "mm" in units else 0.005

    matched_gt: set[str] = set()
    matched_extracted_idx: set[int] = set()
    for gi, g in enumerate(gt_dims):
        for ei, e in enumerate(extracted):
            if ei in matched_extracted_idx:
                continue
            if _match_dim(e, g, value_eps, tol_eps):
                matched_gt.add(g["id"])
                matched_extracted_idx.add(ei)
                break

    tp = len(matched_gt)
    fn = len(gt_dims) - tp
    fp = len(extracted) - len(matched_extracted_idx)
    precision = tp / max(1, tp + fp)
    recall = tp / max(1, tp + fn)
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    return {
        "workflow": "dims_ocr",
        "drawing": drawing_name,
        "model": model,
        "temperature": temperature,
        "ocr_item_count": len(ocr_items),
        "metadata": meta,
        "raw_output": raw,
        "score": {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "tp_count": tp,
            "fp_count": fp,
            "fn_count": fn,
            "matched_ids": sorted(matched_gt),
            "missed_ids": sorted(set(g["id"] for g in gt_dims) - matched_gt),
            "extracted_count": len(extracted),
            "gt_count": len(gt_dims),
        },
    }




# ---------------------- PDF extraction (native + OCR fallback) ----------- #

PDF_TEXT_CACHE = DATA / "docs" / "text_cache"
PDF_TEXT_CACHE.mkdir(parents=True, exist_ok=True)


# ---------------------- dims_vlm (direct image → vision model) ------------ #

def _image_to_data_url(path: Path) -> str:
    """Read a JPEG/PNG image and return an OpenAI-compatible data: URL."""
    import base64
    suffix = path.suffix.lower().lstrip(".")
    mime = "image/jpeg" if suffix in ("jpg", "jpeg") else f"image/{suffix or 'jpeg'}"
    b64 = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{b64}"


def call_vlm(client, model: str, system_prompt: str, user_text: str, image_paths: Path | list[Path], temperature: float) -> tuple[str, dict]:
    """Vision-model variant of call_model. Sends the system prompt as text and the
    user prompt as a (text + image_url[, image_url, ...]) multipart message. Accepts
    either a single Path (single-image, e.g. dims_vlm) or a list of Paths (multi-image,
    e.g. docs_vlm with one image per PDF page). Same retry-on-no-JSON semantics as call_model.

    Transient HTTP errors (429, 502, 503) are retried up to MAX_TRANSIENT_RETRIES
    times with exponential backoff before propagating.

    Captures provider response headers when available."""
    metadata: dict = {"attempts": []}
    if isinstance(image_paths, Path):
        image_paths = [image_paths]
    user_content: list[dict] = [{"type": "text", "text": user_text}]
    for p in image_paths:
        user_content.append({"type": "image_url", "image_url": {"url": _image_to_data_url(p)}})
    metadata["image_count"] = len(image_paths)
    extra_kwargs = _model_extra_kwargs(model)

    create_kwargs = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content},
        ],
        "temperature": temperature,
        "response_format": {"type": "json_object"},
        "timeout": RUN_TIMEOUT_S,
        **extra_kwargs,
    }

    try:
        t0 = time.time()
        with hard_timeout(RUN_TIMEOUT_S * 1.2):
            resp, headers, retries = _create_with_retry(client, create_kwargs)
        elapsed = time.time() - t0
        if headers:
            metadata["response_headers"] = headers
        if retries:
            metadata["transient_retries"] = retries
        content = resp.choices[0].message.content
        if content is not None and content.strip():
            metadata["attempts"].append({"mode": "json_object", "elapsed_s": round(elapsed, 2), "ok": True})
            return content, metadata
        metadata["attempts"].append({"mode": "json_object", "elapsed_s": round(elapsed, 2), "ok": False, "error": "empty content under json_object mode"})
    except Exception as e:
        was_timeout = _is_timeout_error(e)
        metadata["attempts"].append({"mode": "json_object", "ok": False, "error": str(e)[:200], "was_timeout": was_timeout})
        if was_timeout:
            raise

    create_kwargs_plain = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content},
        ],
        "temperature": temperature,
        "timeout": RUN_TIMEOUT_S,
        **extra_kwargs,
    }

    t0 = time.time()
    with hard_timeout(RUN_TIMEOUT_S * 1.2):
        resp, headers, retries = _create_with_retry(client, create_kwargs_plain)
    elapsed = time.time() - t0
    if headers:
        metadata["response_headers"] = headers
    if retries:
        metadata.setdefault("transient_retries", []).extend(retries)
    metadata["attempts"].append({"mode": "plain", "elapsed_s": round(elapsed, 2), "ok": True})
    return resp.choices[0].message.content, metadata


def run_dims_vlm(client: OpenAI, model: str, temperature: float, drawing_name: str) -> dict:
    drawing_path = DATA / "dims" / drawing_name
    gt_path = DATA / "dims" / f"{drawing_path.stem}_ground_truth.json"
    if not drawing_path.is_file() or not gt_path.is_file():
        return {"workflow": "dims_vlm", "drawing": drawing_name, "model": model, "error": "drawing or GT file missing"}

    ground_truth = json.loads(gt_path.read_text(encoding="utf-8"))
    system_prompt = (PROMPTS / "dims_vlm_system.txt").read_text(encoding="utf-8")

    # Per-call salt — appended to the user message to defeat any provider-side
    # response cache that hashes on prompt text rather than honouring the seed.
    nonce = f"req-{int(time.time()*1_000_000) & 0xFFFFFFFF:08x}"
    user_text = (
        f"Engineering drawing image attached.\n"
        f"Title (per spec): {ground_truth.get('title', '?')}\n"
        f"Material: {ground_truth.get('material', '?')}\n"
        f"Units: {ground_truth.get('units', '?')}\n"
        f"Scale: {ground_truth.get('scale', '?')}\n"
        f"General tolerance note: {ground_truth.get('general_tolerance_note', '?')}\n\n"
        f"Look carefully at the image. Engineering drawings have many dimension callouts: "
        f"diameters (Ø…), lengths, radii (R…), angles, chamfers (e.g. 1×45°), threads (e.g. M4), "
        f"surface-finish callouts (Ra…), and GD&T symbols. Every visible number with associated "
        f"feature lines is a dimension and MUST be reported. The drawing typically has at least "
        f"15-30 dimension callouts; if you find fewer than 10, you have missed many — look harder.\n\n"
        f"Return the JSON object as instructed in the system prompt. Do NOT return an empty "
        f"`extracted_dimensions` array — that is never the correct answer for a real engineering drawing.\n\n"
        f"(request id: {nonce})"
    )

    raw, meta = call_vlm(client, model, system_prompt, user_text, drawing_path, temperature)
    try:
        parsed = parse_json_response(raw)
    except Exception as e:
        return {"workflow": "dims_vlm", "drawing": drawing_name, "model": model, "raw_output": raw, "error": f"JSON parse failed: {e}", "metadata": meta}

    extracted = parsed.get("extracted_dimensions", []) if isinstance(parsed, dict) else []
    gt_dims = ground_truth.get("expected_dimensions", [])
    units = (ground_truth.get("units") or "").lower()
    value_eps = 0.05 if "milli" in units or "mm" in units else 0.01
    tol_eps = 0.05 if "milli" in units or "mm" in units else 0.005

    matched_gt: set[str] = set()
    matched_extracted_idx: set[int] = set()
    for g in gt_dims:
        for ei, e in enumerate(extracted):
            if ei in matched_extracted_idx:
                continue
            if _match_dim(e, g, value_eps, tol_eps):
                matched_gt.add(g["id"])
                matched_extracted_idx.add(ei)
                break

    tp = len(matched_gt)
    fn = len(gt_dims) - tp
    fp = len(extracted) - len(matched_extracted_idx)
    precision = tp / max(1, tp + fp)
    recall = tp / max(1, tp + fn)
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    return {
        "workflow": "dims_vlm",
        "drawing": drawing_name,
        "model": model,
        "temperature": temperature,
        "metadata": meta,
        "raw_output": raw,
        "score": {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "tp_count": tp,
            "fp_count": fp,
            "fn_count": fn,
            "matched_ids": sorted(matched_gt),
            "missed_ids": sorted({g["id"] for g in gt_dims} - matched_gt),
            "extracted_count": len(extracted),
            "gt_count": len(gt_dims),
        },
    }


# ---------------------- docs_vlm (direct page images → vision model) ----- #

PDF_PAGE_CACHE = DATA / "docs" / "page_cache"
PDF_PAGE_CACHE.mkdir(parents=True, exist_ok=True)


def _render_pdf_pages_to_pngs(pdf_path: Path, dpi: int = 150) -> list[Path]:
    """Render every page of `pdf_path` to a PNG and cache on disk by stem + page index.
    Returns a list of PNG paths in page order. Uses PyMuPDF.

    Auto-detects an empty / partial cache (e.g. from an interrupted run) and re-renders.
    DPI defaults to 150 — VLMs accept this comfortably and it keeps the data URLs reasonable.
    """
    try:
        import pymupdf  # type: ignore
    except ImportError:
        raise RuntimeError("pymupdf not installed. Run: uv sync")

    doc = pymupdf.open(str(pdf_path))
    n_pages = len(doc)
    cached_paths = [PDF_PAGE_CACHE / f"{pdf_path.stem}_p{i:02d}.png" for i in range(n_pages)]
    if all(p.is_file() and p.stat().st_size > 0 for p in cached_paths) and len(cached_paths) == n_pages:
        return cached_paths

    out: list[Path] = []
    for i, page in enumerate(doc):
        target = PDF_PAGE_CACHE / f"{pdf_path.stem}_p{i:02d}.png"
        pix = page.get_pixmap(dpi=dpi)
        target.write_bytes(pix.tobytes("png"))
        out.append(target)
    return out


def run_docs_vlm(client: OpenAI, model: str, temperature: float, pdf_name: str) -> dict:
    pdf_path = DATA / "docs" / pdf_name
    gt_path = DATA / "docs" / f"{pdf_path.stem}_ground_truth.json"
    if not pdf_path.is_file() or not gt_path.is_file():
        return {"workflow": "docs_vlm", "pdf": pdf_name, "model": model, "error": "PDF or GT file missing"}

    ground_truth = json.loads(gt_path.read_text(encoding="utf-8"))

    try:
        page_images = _render_pdf_pages_to_pngs(pdf_path)
    except Exception as e:
        return {"workflow": "docs_vlm", "pdf": pdf_name, "model": model, "error": f"PDF page render failed: {e}"}

    if not page_images:
        return {"workflow": "docs_vlm", "pdf": pdf_name, "model": model, "error": "PDF has no renderable pages"}

    system_prompt = (PROMPTS / "docs_vlm_system.txt").read_text(encoding="utf-8")
    nonce = f"req-{int(time.time()*1_000_000) & 0xFFFFFFFF:08x}"
    user_text = (
        f"Document file name (for context): {pdf_name}\n"
        f"Number of page images attached: {len(page_images)}\n\n"
        f"Read the page image(s) carefully and extract every structured entity present "
        f"in this document — people, organisations, addresses, emails, phones, dates, "
        f"monetary values, identifiers, legal references, deadlines, etc. A real document "
        f"typically has 10+ entities; do NOT return an empty `entities` array.\n\n"
        f"Return the JSON object as instructed in the system prompt.\n\n"
        f"(request id: {nonce})"
    )

    raw, meta = call_vlm(client, model, system_prompt, user_text, page_images, temperature)
    meta["page_count"] = len(page_images)

    try:
        parsed = parse_json_response(raw)
    except Exception as e:
        return {"workflow": "docs_vlm", "pdf": pdf_name, "model": model, "raw_output": raw, "error": f"JSON parse failed: {e}", "metadata": meta}

    extracted = parsed.get("entities", []) if isinstance(parsed, dict) else []

    def _entity_text(e: dict) -> str:
        return ((e.get("kind", "") or "") + " :: " + str(e.get("value", "") or "") + " :: " + (e.get("context", "") or "")).lower()

    entity_texts = [_entity_text(e) for e in extracted if isinstance(e, dict)]

    matched: set[str] = set()
    matched_extracted_idx: set[int] = set()
    for g in ground_truth.get("expected_entities", []):
        kws = [k.lower() for k in g.get("match_keywords_any_of", [])]
        for i, t in enumerate(entity_texts):
            if i in matched_extracted_idx:
                continue
            if any(k in t for k in kws):
                matched.add(g["id"])
                matched_extracted_idx.add(i)
                break

    n_total = len(ground_truth.get("expected_entities", []))
    tp = len(matched)
    fn = n_total - tp
    extracted_count = len(extracted)
    fp_estimate = max(0, extracted_count - tp)
    precision = tp / max(1, tp + fp_estimate)
    recall = tp / max(1, n_total)
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    return {
        "workflow": "docs_vlm",
        "pdf": pdf_name,
        "model": model,
        "temperature": temperature,
        "page_count": len(page_images),
        "metadata": meta,
        "raw_output": raw,
        "score": {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "tp_count": tp,
            "fp_count_estimate": fp_estimate,
            "fn_count": fn,
            "matched_ids": sorted(matched),
            "missed_ids": sorted({g["id"] for g in ground_truth.get("expected_entities", [])} - matched),
            "extracted_count": extracted_count,
            "gt_count": n_total,
        },
    }


def _extract_pdf_text(pdf_path: Path) -> tuple[str, str]:
    """Return (text, mode) where mode is 'native' or 'ocr'. Cached on disk by stem.

    Auto-detects corrupted / empty caches (e.g. from a previously interrupted run)
    and regenerates them. A valid cache is a dict with non-empty 'text' (>= 50 chars).
    """
    cache = PDF_TEXT_CACHE / f"{pdf_path.stem}.json"
    if cache.is_file():
        try:
            d = json.loads(cache.read_text(encoding="utf-8"))
            if isinstance(d, dict) and isinstance(d.get("text"), str) and len(d["text"]) >= 50:
                return d["text"], d.get("mode", "native")
            print(f"  ⚠ PDF text cache for {pdf_path.name} is empty / invalid — regenerating.")
        except Exception as e:
            print(f"  ⚠ PDF text cache for {pdf_path.name} unreadable ({e}) — regenerating.")
        try:
            cache.unlink()
        except Exception:
            pass

    try:
        import pymupdf  # type: ignore
    except ImportError:
        raise RuntimeError("pymupdf not installed. Run: uv sync")

    doc = pymupdf.open(str(pdf_path))
    pages_native: list[str] = []
    for p in doc:
        pages_native.append(p.get_text("text"))
    native_text = "\n\n".join(pages_native).strip()

    if len(native_text) >= 50 * len(doc):
        # ~50 chars per page is the threshold for "this PDF has a real text layer".
        cache.write_text(json.dumps({"text": native_text, "mode": "native"}, ensure_ascii=False), encoding="utf-8")
        return native_text, "native"

    # Fallback: render each page and OCR with EasyOCR.
    reader = get_ocr_reader()
    pages_ocr: list[str] = []
    for i, p in enumerate(doc):
        pix = p.get_pixmap(dpi=200)
        png_bytes = pix.tobytes("png")
        ocr_lines = reader.readtext(png_bytes, detail=0)
        pages_ocr.append("\n".join(ocr_lines))
    ocr_text = "\n\n--- page break ---\n\n".join(pages_ocr).strip()

    cache.write_text(json.dumps({"text": ocr_text, "mode": "ocr"}, ensure_ascii=False), encoding="utf-8")
    return ocr_text, "ocr"


def run_docs(client: OpenAI, model: str, temperature: float, pdf_name: str) -> dict:
    pdf_path = DATA / "docs" / pdf_name
    gt_path = DATA / "docs" / f"{pdf_path.stem}_ground_truth.json"
    if not pdf_path.is_file() or not gt_path.is_file():
        return {"workflow": "docs", "pdf": pdf_name, "model": model, "error": "PDF or GT file missing"}

    ground_truth = json.loads(gt_path.read_text(encoding="utf-8"))

    try:
        text, mode = _extract_pdf_text(pdf_path)
    except Exception as e:
        return {"workflow": "docs", "pdf": pdf_name, "model": model, "error": f"PDF extraction failed: {e}"}

    system_prompt = (PROMPTS / "docs_system.txt").read_text(encoding="utf-8")
    user_prompt = (
        f"Document file name (for context): {pdf_name}\n"
        f"Extraction mode (for context): {mode}\n\n"
        f"DOCUMENT TEXT:\n{'='*60}\n{text}\n{'='*60}\n\n"
        f"Extract every structured entity present in this document and return the JSON object as instructed."
    )

    raw, meta = call_model(client, model, system_prompt, user_prompt, temperature)
    try:
        parsed = parse_json_response(raw)
    except Exception as e:
        return {"workflow": "docs", "pdf": pdf_name, "model": model, "raw_output": raw, "error": f"JSON parse failed: {e}", "metadata": meta}

    extracted = parsed.get("entities", []) if isinstance(parsed, dict) else []

    def _entity_text(e: dict) -> str:
        return ((e.get("kind", "") or "") + " :: " + str(e.get("value", "") or "") + " :: " + (e.get("context", "") or "")).lower()

    entity_texts = [_entity_text(e) for e in extracted if isinstance(e, dict)]

    # Per-entity matching: for each GT entry, look for at least one extracted entity whose text
    # contains one of the GT keywords. This avoids the cross-entity bleed where e.g. an
    # address mentioning "Germany" would otherwise match a GT "language: German" entry.
    matched: set[str] = set()
    matched_extracted_idx: set[int] = set()
    for g in ground_truth.get("expected_entities", []):
        kws = [k.lower() for k in g.get("match_keywords_any_of", [])]
        for i, t in enumerate(entity_texts):
            if i in matched_extracted_idx:
                continue
            if any(k in t for k in kws):
                matched.add(g["id"])
                matched_extracted_idx.add(i)
                break

    n_total = len(ground_truth.get("expected_entities", []))
    tp = len(matched)
    fn = n_total - tp
    extracted_count = len(extracted)
    fp_estimate = max(0, extracted_count - tp)
    precision = tp / max(1, tp + fp_estimate)
    recall = tp / max(1, n_total)
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    return {
        "workflow": "docs",
        "pdf": pdf_name,
        "model": model,
        "temperature": temperature,
        "extraction_mode": mode,
        "metadata": meta,
        "raw_output": raw,
        "score": {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "tp_count": tp,
            "fp_count_estimate": fp_estimate,
            "fn_count": fn,
            "matched_ids": sorted(matched),
            "missed_ids": sorted({g["id"] for g in ground_truth.get("expected_entities", [])} - matched),
            "extracted_count": extracted_count,
            "gt_count": n_total,
        },
    }


# ---------------------- schedule_read (spreadsheet read) ----------------- #

def _read_xlsx_as_table(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: uv sync")
    wb = load_workbook(str(path), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        out.append({h: ("" if v is None else str(v)) for h, v in zip(headers, row)})
    return out


def run_schedule_read(client: OpenAI, model: str, temperature: float) -> dict:
    sheet_path = DATA / "schedule_read" / "timeline.xlsx"
    gt_path = DATA / "schedule_read" / "ground_truth.json"
    if not sheet_path.is_file() or not gt_path.is_file():
        return {"workflow": "schedule_read", "model": model, "error": "spreadsheet or GT missing"}

    table = _read_xlsx_as_table(sheet_path)
    ground_truth = json.loads(gt_path.read_text(encoding="utf-8"))

    system_prompt = (PROMPTS / "schedule_read_system.txt").read_text(encoding="utf-8")

    if not table:
        return {"workflow": "schedule_read", "model": model, "error": "spreadsheet is empty"}
    headers = list(table[0].keys())
    text = " | ".join(headers) + "\n"
    for r in table:
        text += " | ".join(str(r.get(h, "")) for h in headers) + "\n"

    user_prompt = (
        "PROJECT GANTT TIMELINE (rendered as a pipe-delimited table):\n"
        f"{text}\n"
        "Review the timeline and return the JSON object as instructed."
    )

    raw, meta = call_model(client, model, system_prompt, user_prompt, temperature)
    try:
        parsed = parse_json_response(raw)
    except Exception as e:
        return {"workflow": "schedule_read", "model": model, "raw_output": raw, "error": f"JSON parse failed: {e}", "metadata": meta}

    flagged = parsed.get("issues", []) if isinstance(parsed, dict) else []
    flagged_blob = [
        ((f.get("kind", "") or "") + " " + (f.get("part_id", "") or "") + " " +
         " ".join(f.get("stages_involved", []) or []) + " " + (f.get("summary", "") or "")).lower()
        for f in flagged if isinstance(f, dict)
    ]

    matched: set[str] = set()
    matched_flag_indices: set[int] = set()
    # Also pull each LLM flag's explicit part_id field for a stricter alignment check.
    flag_part_ids = [(f.get("part_id", "") or "").lower() for f in flagged if isinstance(f, dict)]
    for g in ground_truth.get("expected_issues", []):
        kws = [k.lower() for k in g.get("match_keywords_any_of", [])]
        gt_part_ids = [g.get("part_id", "").lower()] + [p.lower() for p in g.get("involved_parts", []) or []]
        gt_part_ids = [p for p in gt_part_ids if p]
        for i, blob in enumerate(flagged_blob):
            if i in matched_flag_indices:
                continue
            hits = sum(1 for k in kws if k in blob)
            # Strict alignment: if the LLM flag has an explicit part_id, it must align with GT.
            llm_pid = flag_part_ids[i] if i < len(flag_part_ids) else ""
            if gt_part_ids and llm_pid and llm_pid not in gt_part_ids:
                continue
            if hits >= 2 or (hits >= 1 and any(p in blob for p in gt_part_ids)):
                matched.add(g["id"])
                matched_flag_indices.add(i)
                break

    n_total = len(ground_truth.get("expected_issues", []))
    tp = len(matched)
    fp = len(flagged) - len(matched_flag_indices)
    fn = n_total - tp
    precision = tp / max(1, tp + fp)
    recall = tp / max(1, n_total)
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    return {
        "workflow": "schedule_read",
        "model": model,
        "temperature": temperature,
        "metadata": meta,
        "raw_output": raw,
        "score": {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "tp_count": tp,
            "fp_count": fp,
            "fn_count": fn,
            "matched_ids": sorted(matched),
            "missed_ids": sorted({g["id"] for g in ground_truth.get("expected_issues", [])} - matched),
            "flagged_count": len(flagged),
            "gt_count": n_total,
        },
    }


# ---------------------- schedule_write (spreadsheet write) --------------- #

def run_schedule_write(client: OpenAI, model: str, temperature: float) -> dict:
    base_path = DATA / "schedule_write" / "timeline_baseline.xlsx"
    scenario_path = DATA / "schedule_write" / "scenario.txt"
    gt_path = DATA / "schedule_write" / "ground_truth.json"
    if not all(p.is_file() for p in (base_path, scenario_path, gt_path)):
        return {"workflow": "schedule_write", "model": model, "error": "baseline / scenario / GT file missing"}

    table = _read_xlsx_as_table(base_path)
    if not table:
        return {"workflow": "schedule_write", "model": model, "error": "baseline spreadsheet empty"}
    scenario = scenario_path.read_text(encoding="utf-8")
    ground_truth = json.loads(gt_path.read_text(encoding="utf-8"))

    headers = list(table[0].keys())
    text = " | ".join(headers) + "\n"
    for r in table:
        text += " | ".join(str(r.get(h, "")) for h in headers) + "\n"

    system_prompt = (PROMPTS / "schedule_write_system.txt").read_text(encoding="utf-8")
    user_prompt = (
        "CURRENT GANTT TIMELINE (pipe-delimited):\n"
        f"{text}\n\n"
        "SCENARIO:\n"
        f"{scenario}\n\n"
        "Produce the JSON edit list as instructed."
    )

    raw, meta = call_model(client, model, system_prompt, user_prompt, temperature)
    try:
        parsed = parse_json_response(raw)
    except Exception as e:
        return {"workflow": "schedule_write", "model": model, "raw_output": raw, "error": f"JSON parse failed: {e}", "metadata": meta}

    edits = parsed.get("edits", []) if isinstance(parsed, dict) else []

    def norm(s: Any) -> str:
        return re.sub(r"\s+", " ", str(s or "").lower().strip())

    expected = ground_truth.get("expected_edits", [])
    matched: set[str] = set()
    matched_edit_indices: set[int] = set()
    for ge in expected:
        for ei, e in enumerate(edits):
            if ei in matched_edit_indices or not isinstance(e, dict):
                continue
            if norm(e.get("part_id")) != norm(ge["part_id"]):
                continue
            if norm(ge["stage"]) not in norm(e.get("stage")) and norm(e.get("stage")) not in norm(ge["stage"]):
                continue
            if norm(e.get("column")) != norm(ge["column"]):
                continue
            new_val = norm(e.get("new_value"))
            ok = False
            if "expected_value" in ge:
                ok = norm(ge["expected_value"]) == new_val or norm(ge["expected_value"]) in new_val
                if not ok and "expected_value_alternatives" in ge:
                    ok = any(norm(alt) == new_val or norm(alt) in new_val for alt in ge["expected_value_alternatives"])
            elif "expected_value_keywords_any_of" in ge:
                ok = any(k.lower() in new_val for k in ge["expected_value_keywords_any_of"])
            if ok:
                matched.add(ge["id"])
                matched_edit_indices.add(ei)
                break

    protected = ground_truth.get("must_not_modify", {}).get("rows_protected", [])
    forbidden_violations = []
    for ei, e in enumerate(edits):
        if ei in matched_edit_indices or not isinstance(e, dict):
            continue
        for pr in protected:
            if norm(e.get("part_id")) == norm(pr["part_id"]) and norm(pr["stage"]) in norm(e.get("stage")):
                forbidden_violations.append({"part_id": e.get("part_id"), "stage": e.get("stage"), "column": e.get("column")})
                break

    n_total = len(expected)
    tp = len(matched)
    fp = len(edits) - len(matched_edit_indices)
    fn = n_total - tp
    precision = tp / max(1, tp + fp)
    recall = tp / max(1, n_total)
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    return {
        "workflow": "schedule_write",
        "model": model,
        "temperature": temperature,
        "metadata": meta,
        "raw_output": raw,
        "score": {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "tp_count": tp,
            "fp_count": fp,
            "fn_count": fn,
            "forbidden_row_violations": forbidden_violations,
            "matched_ids": sorted(matched),
            "missed_ids": sorted({g["id"] for g in expected} - matched),
            "emitted_edit_count": len(edits),
            "gt_edit_count": n_total,
        },
    }


# ----------------------------- aggregator --------------------------------- #

_INFRA_FAILURE_PATTERNS = [
    "402", "404", "429", "502", "503", "520",
    "timed out", "timeout", "circuit-breaker",
    "connection", "rate limit", "credit", "payment", "skipped:",
]


def _is_infra_failure(error_str: str) -> bool:
    """Classify a run error as infrastructure (provider) vs capability (model).

    Infrastructure failures: credit exhaustion, routing errors, timeouts,
    rate limits, connection issues, circuit-breaker skips.
    Capability failures: JSON parse errors, empty content, or anything else
    that implies the model got a chance to respond but produced unusable output.
    """
    s = (error_str or "").lower()
    return any(p in s for p in _INFRA_FAILURE_PATTERNS)


def _bootstrap_ci(values: list[float], n_boot: int = 10000, ci: float = 0.95) -> list[float]:
    """Bootstrap 95% CI for the mean. Returns [lo, hi]."""
    if len(values) < 2:
        m = values[0] if values else 0.0
        return [round(m, 4), round(m, 4)]
    means = [statistics.fmean(_random.choices(values, k=len(values))) for _ in range(n_boot)]
    means.sort()
    lo = means[int((1 - ci) / 2 * n_boot)]
    hi = means[int((1 + ci) / 2 * n_boot)]
    return [round(lo, 4), round(hi, 4)]


def aggregate(runs: list[dict]) -> dict:
    """Compute mean / std / CI / min / max for precision, recall, F1, latency.

    Separates infrastructure failures (provider errors) from capability failures
    (model produced unusable output). Reports two score sets:
      - scores_all_runs:        all failures score 0.0 (backward-compatible)
      - scores_successful_only: P/R/F1 computed only over successful runs

    Plus a reliability block showing success rate and failure breakdown.
    """
    successful = [r for r in runs if "error" not in r]
    failed = [r for r in runs if "error" in r]
    infra_failures = [r for r in failed if _is_infra_failure(r.get("error", ""))]
    capability_failures = [r for r in failed if not _is_infra_failure(r.get("error", ""))]

    def _stats(values: list[float]) -> dict:
        if not values:
            return {"mean": None, "std": None, "min": None, "max": None, "ci_95": None}
        return {
            "mean": round(statistics.fmean(values), 4),
            "std": round(statistics.pstdev(values), 4) if len(values) > 1 else 0.0,
            "min": round(min(values), 4),
            "max": round(max(values), 4),
            "ci_95": _bootstrap_ci(values),
        }

    # scores_all_runs: failed runs score 0.0 (backward-compatible).
    all_precisions = [r["score"]["precision"] for r in successful] + [0.0] * len(failed)
    all_recalls    = [r["score"]["recall"]    for r in successful] + [0.0] * len(failed)
    all_f1s        = [r["score"]["f1"]        for r in successful] + [0.0] * len(failed)

    # scores_successful_only: only runs that produced valid output.
    succ_precisions = [r["score"]["precision"] for r in successful]
    succ_recalls    = [r["score"]["recall"]    for r in successful]
    succ_f1s        = [r["score"]["f1"]        for r in successful]

    latencies: list[float] = []
    for r in successful:
        attempts = r.get("metadata", {}).get("attempts", [])
        for a in attempts:
            if a.get("ok"):
                latencies.append(float(a.get("elapsed_s", 0)))
                break

    total = len(runs)
    return {
        "reliability": {
            "total_runs": total,
            "successful_runs": len(successful),
            "infra_failures": len(infra_failures),
            "capability_failures": len(capability_failures),
            "success_rate": round(len(successful) / max(1, total), 4),
        },
        "scores_all_runs": {
            "precision": _stats(all_precisions),
            "recall": _stats(all_recalls),
            "f1": _stats(all_f1s),
        },
        "scores_successful_only": {
            "precision": _stats(succ_precisions),
            "recall": _stats(succ_recalls),
            "f1": _stats(succ_f1s),
        },
        # Backward-compatible top-level aliases (point to scores_all_runs).
        "successful_runs": len(successful),
        "failed_runs": len(failed),
        "errors": [r.get("error") for r in failed][:5],
        "precision": _stats(all_precisions),
        "recall": _stats(all_recalls),
        "f1": _stats(all_f1s),
        "latency_s": _stats(latencies),
    }


# ----------------------------- summary printers --------------------------- #

def fmt_pm(stat: dict, places: int = 3) -> str:
    if stat is None or stat.get("mean") is None:
        return "—"
    return f"{stat['mean']:.{places}f} ± {stat['std']:.{places}f}"


def fmt_ci(stat: dict, places: int = 3) -> str:
    """Format a stat dict as 'mean [lo, hi]' using 95% CI."""
    if stat is None or stat.get("mean") is None:
        return "—"
    ci = stat.get("ci_95")
    if ci:
        return f"{stat['mean']:.{places}f} [{ci[0]:.{places}f}, {ci[1]:.{places}f}]"
    return f"{stat['mean']:.{places}f}"


def _reliability_str(agg: dict) -> str:
    """Format reliability as 'ok/total' string."""
    rel = agg.get("reliability", {})
    total = rel.get("total_runs", agg.get("successful_runs", 0) + agg.get("failed_runs", 0))
    ok = rel.get("successful_runs", agg.get("successful_runs", 0))
    return f"{ok}/{total}"


def render_summary_md(report: dict) -> str:
    lines: list[str] = []
    lines.append("# AGEN-OPS-6 benchmark summary\n")
    lines.append(f"Generated: {report['timestamp']}  ")
    lines.append(f"Total runs requested per (workflow, model[, drawing]) combination: {report['runs_per_combo']}\n")

    lines.append("## Headline (successful runs only)\n")
    lines.append("| Workflow | Best model (by F1) | F1 [95% CI] | Runs ok | Latency mean (s) |")
    lines.append("|---|---|---|---|---|")
    for wf_key, wf_block in report["workflows"].items():
        best = None
        for combo in wf_block["combos"]:
            agg = combo["aggregate"]
            # Use successful-only F1 as primary ranking metric.
            succ_f1 = (agg.get("scores_successful_only", {}).get("f1") or agg.get("f1") or {}).get("mean")
            if succ_f1 is None:
                continue
            if best is None:
                best = combo
            else:
                best_f1 = (best["aggregate"].get("scores_successful_only", {}).get("f1") or best["aggregate"].get("f1") or {}).get("mean", 0)
                if succ_f1 > best_f1:
                    best = combo
        if best is None:
            continue
        agg = best["aggregate"]
        label = best["model"] + (f" / {best.get('drawing') or best.get('pdf')}" if (best.get('drawing') or best.get('pdf')) else "")
        succ_f1_stat = agg.get("scores_successful_only", {}).get("f1") or agg.get("f1") or {}
        lines.append(
            f"| {wf_block['name']} | {label} | {fmt_ci(succ_f1_stat)} | "
            f"{_reliability_str(agg)} | "
            f"{(agg['latency_s']['mean'] if agg['latency_s']['mean'] is not None else 0):.1f} |"
        )
    lines.append("")

    for wf_key, wf_block in report["workflows"].items():
        lines.append(f"## {wf_block['name']}\n")
        lines.append(wf_block.get("description", "") + "\n")
        lines.append("| Model | Drawing/PDF | Runs ok | F1 all [CI] | F1 succ [CI] | P succ | R succ | Latency (s) |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for combo in wf_block["combos"]:
            agg = combo["aggregate"]
            item_label = combo.get("drawing") or combo.get("pdf") or "—"
            all_f1 = agg.get("scores_all_runs", {}).get("f1") or agg.get("f1") or {}
            succ_f1 = agg.get("scores_successful_only", {}).get("f1") or agg.get("f1") or {}
            succ_p = agg.get("scores_successful_only", {}).get("precision") or agg.get("precision") or {}
            succ_r = agg.get("scores_successful_only", {}).get("recall") or agg.get("recall") or {}
            lines.append(
                f"| {combo['model']} | {item_label} | "
                f"{_reliability_str(agg)} | "
                f"{fmt_ci(all_f1)} | {fmt_ci(succ_f1)} | "
                f"{fmt_pm(succ_p)} | {fmt_pm(succ_r)} | "
                f"{(agg['latency_s']['mean'] if agg['latency_s']['mean'] is not None else 0):.1f} |"
            )
        lines.append("")

    return "\n".join(lines)


def print_stdout_summary(report: dict) -> None:
    print()
    print("=" * 78)
    print("EXPERIMENT SUMMARY")
    print("=" * 78)
    for wf_key, wf_block in report["workflows"].items():
        print()
        print(f"-- {wf_block['name']} --")
        for combo in wf_block["combos"]:
            agg = combo["aggregate"]
            label = f"{combo['model']:48s}"
            item_label = combo.get("drawing") or combo.get("pdf")
            if item_label:
                label = label[:35] + f" / {item_label[:24]:25s}"
            ok = _reliability_str(agg)
            # Show successful-only F1 as primary, with reliability qualifier.
            succ_f1_stat = agg.get("scores_successful_only", {}).get("f1") or agg.get("f1") or {}
            f1m = succ_f1_stat.get("mean", 0) or 0
            ci = succ_f1_stat.get("ci_95")
            ci_str = f" [{ci[0]:.3f},{ci[1]:.3f}]" if ci else ""
            p = fmt_pm(agg.get("scores_successful_only", {}).get("precision") or agg.get("precision", {}))
            r = fmt_pm(agg.get("scores_successful_only", {}).get("recall") or agg.get("recall", {}))
            lat = agg["latency_s"]["mean"] if agg["latency_s"]["mean"] is not None else 0
            print(f"  {label:60s}  runs={ok:5s}  P={p}  R={r}  F1={f1m:.3f}{ci_str}  lat={lat:.1f}s")
    print()


# ----------------------------- main loop ---------------------------------- #

def short_model(m: str) -> str:
    return m.replace("/", "_").replace(":", "_")


def main() -> None:
    parser = argparse.ArgumentParser(description="AGEN-OPS-6 benchmark runner — multi-backend, multi-model, multi-run, statistical aggregation.")
    parser.add_argument("--workflow", choices=["compliance", "dims", "dims_ocr", "dims_vlm", "docs", "docs_vlm", "schedule_read", "schedule_write", "all"], default="all")
    parser.add_argument("--models", nargs="+", default=None,
                        help="Model ids to compare. If omitted, uses per-workflow defaults: "
                             "text workflows → full Tensorix ladder; dims_ocr → vision ladder; "
                             "dims_vlm / docs_vlm → OpenRouter VLM ladder.")
    parser.add_argument("--runs", type=int, default=10, help="Runs per (workflow, model[, drawing]) combination.")
    parser.add_argument("--temperature", type=float, default=0.1)
    parser.add_argument("--drawings", nargs="+", default=["technical-drawing-1.jpg", "technical-drawing-2.jpg", "technical-drawing-3.jpg", "technical-drawing-4.jpg", "technical-drawing-5.jpg", "technical-drawing-6.jpg"], help="Drawings for dims_ocr / dims_vlm.")
    parser.add_argument("--pdfs", nargs="+", default=["llm_finetuning_report.pdf", "llm_finetuning_report_scanned.pdf", "llm_finetuning_report_image.pdf"], help="PDFs for docs / docs_vlm.")
    parser.add_argument("--no-extract", action="store_true", help="Skip the dims_ocr workflow (which needs EasyOCR).")
    parser.add_argument("--no-save", action="store_true")
    parser.add_argument("--workers", type=int, default=1,
                        help="Number of (model, item) combos to run in parallel within each workflow (default: 1 = sequential). "
                             "The outer workflow loop stays sequential; only combos within a workflow are parallelised.")
    parser.add_argument("--delay", type=float, default=0.5,
                        help="Seconds to sleep before each LLM API call to avoid provider rate limits (default: 0.5).")
    parser.add_argument("--clear-cache", action="store_true",
                        help="Delete OCR + PDF text + PDF page caches at start of run, forcing re-OCR / re-parse / re-render. "
                             "Use after an interrupted run if scores look suspicious (P=R=0 across the board often means a corrupted cache).")
    parser.add_argument("--check-credits", action="store_true",
                        help="Check OpenRouter credit balance before running. Prints a warning if balance is low.")
    args = parser.parse_args()

    if args.check_credits:
        or_key = os.environ.get("OPENROUTER_API_KEY")
        if or_key:
            try:
                import urllib.request
                req = urllib.request.Request(
                    "https://openrouter.ai/api/v1/auth/key",
                    headers={"Authorization": f"Bearer {or_key}"},
                )
                with urllib.request.urlopen(req, timeout=10) as resp:
                    info = json.loads(resp.read().decode())
                    balance = info.get("data", {}).get("limit_remaining")
                    usage = info.get("data", {}).get("usage")
                    label = info.get("data", {}).get("label", "?")
                    print(f"OpenRouter credit check:")
                    print(f"  Key label:  {label}")
                    if balance is not None:
                        print(f"  Remaining:  ${balance:.2f}")
                        if balance < 1.0:
                            print(f"  ⚠ WARNING: Balance is low (${balance:.2f}). Top up before a full run.")
                    elif usage is not None:
                        print(f"  Usage:      ${usage:.4f}")
                        print(f"  (No hard limit set — usage is pay-as-you-go)")
                    else:
                        print(f"  (Could not determine balance from API response)")
                    print()
            except Exception as e:
                print(f"OpenRouter credit check failed: {e}", file=sys.stderr)
                print()
        else:
            print("OpenRouter credit check skipped (OPENROUTER_API_KEY not set).")
            print()

    if args.clear_cache:
        cleared = []
        for cache_dir in (OCR_CACHE, PDF_TEXT_CACHE, PDF_PAGE_CACHE):
            if cache_dir.is_dir():
                for f in cache_dir.iterdir():
                    if f.is_file():
                        try:
                            f.unlink()
                            cleared.append(str(f.relative_to(ROOT)))
                        except Exception as e:
                            print(f"  could not delete {f}: {e}", file=sys.stderr)
        if cleared:
            print(f"Cleared {len(cleared)} cache file(s): " + ", ".join(cleared[:5]) + (" ..." if len(cleared) > 5 else ""))
        else:
            print("No cache files to clear.")
        print()

    workflows: list[str]
    if args.workflow == "all":
        # dims_ocr dropped from default (OCR bottleneck; see docs/FIRST_RUN_INSIGHTS.md).
        # Still runnable via --workflow dims_ocr for appendix / replication.
        # docs_vlm dropped from default (0/10 in first run due to provider failures;
        # docs with OCR fallback is the primary PDF extraction path).
        workflows = ["compliance", "dims", "docs", "schedule_read", "schedule_write"]
        # dims_vlm is the primary vision workflow — include if OpenRouter or
        # any VLM-capable backend is configured.
        if os.environ.get("OPENROUTER_API_KEY"):
            dims_idx = workflows.index("dims")
            workflows.insert(dims_idx + 1, "dims_vlm")
    else:
        workflows = [args.workflow]

    def models_for(wf: str) -> list[str]:
        if args.models:
            return list(args.models)
        return list(DEFAULT_MODELS_PER_WORKFLOW.get(wf, DEFAULT_MODELS))

    print(f"Workflows:    {', '.join(workflows)}")
    if args.models:
        print(f"Models (override applies to every workflow): {', '.join(args.models)}")
    else:
        for w in workflows:
            print(f"  {w:14s} models: {', '.join(models_for(w))}")
    print(f"Runs per combo: {args.runs}")
    print(f"Workers:      {args.workers}  (parallel combos per workflow)")
    print(f"Delay:        {args.delay}s  (sleep before each LLM call)")
    print(f"Temperature:  {args.temperature}")
    print(f"Tensorix URL: {os.environ.get('TENSORIX_BASE_URL', '(not configured)')}")
    if os.environ.get('OPENROUTER_API_KEY'):
        print(f"OpenRouter:   {os.environ.get('OPENROUTER_BASE_URL', 'https://openrouter.ai/api/v1')}")
    else:
        print("OpenRouter:   (not configured — OPENROUTER_API_KEY missing; VLMs + anchors will fail)")
    print()

    pool = make_client_pool()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

    def _combo_count(wf: str) -> int:
        n = len(models_for(wf))
        if wf in ("dims_ocr", "dims_vlm"):
            return n * len(args.drawings)
        if wf in ("docs", "docs_vlm"):
            return n * len(args.pdfs)
        return n

    total_combos = sum(_combo_count(w) for w in workflows)
    total_runs_planned = total_combos * args.runs
    # Thread-safe progress counters (list-of-one so inner functions can mutate).
    counters = {"runs": 0, "combos": 0}
    counter_lock = threading.Lock()
    file_lock = threading.Lock()
    print_lock = threading.Lock()
    print(f"Plan: {total_combos} combos x {args.runs} runs = {total_runs_planned} LLM calls\n")

    report: dict = {
        "timestamp": timestamp,
        "models_per_workflow": {w: models_for(w) for w in workflows},
        "runs_per_combo": args.runs,
        "temperature": args.temperature,
        "workflows": {},
    }

    workflow_meta = {
        "compliance":     {"name": "compliance — Material certificate compliance",
                           "description": "Synthetic spec + cert with 5 deliberate deviations + 1 borderline trap."},
        "dims":           {"name": "dims — Drawing-vs-3D dimension comparison reasoning",
                           "description": "25 dimensions; 9 must flag at the 10% tolerance-band rule."},
        "dims_ocr":       {"name": "dims_ocr — Vision extraction (EasyOCR + LLM hybrid)",
                           "description": "Local EasyOCR + LLM consolidates dim+tolerance table."},
        "dims_vlm":       {"name": "dims_vlm — Direct vision extraction (VLM, no OCR)",
                           "description": "Drawing image sent directly to a vision-language model. Tests off-the-shelf VLM upper bound before fine-tuning."},
        "docs":           {"name": "docs — Structured extraction with native + OCR fallback",
                           "description": "PyMuPDF + EasyOCR fallback; LLM consolidates entities."},
        "docs_vlm":       {"name": "docs_vlm — Direct vision extraction (VLM, no OCR)",
                           "description": "PDF pages rendered to PNG and sent directly to a vision-language model. Tests off-the-shelf VLM upper bound on mixed native + scanned docs."},
        "schedule_read":  {"name": "schedule_read — Spreadsheet read: detect scheduling issues",
                           "description": "Project Gantt timeline with 4 deliberately-injected issues."},
        "schedule_write": {"name": "schedule_write — Spreadsheet write: propose cascade-update edits",
                           "description": "Clean Gantt baseline plus a supplier-delay email scenario."},
    }

    def _run_combo(wf: str, model: str, item: str | None,
                   temperature: float, n_runs: int, delay: float,
                   pool: ClientPool, timestamp: str, no_save: bool,
                   total_combos: int, total_runs_planned: int,
                   counters: dict, counter_lock: threading.Lock,
                   file_lock: threading.Lock, print_lock: threading.Lock) -> dict:
        """Execute N runs for a single (workflow, model, item) combo. Thread-safe."""
        label = f"{wf}/{model}" + (f"/{item}" if item else "")
        with counter_lock:
            counters["combos"] += 1
            combo_num = counters["combos"]
        with print_lock:
            print(f"[combo {combo_num}/{total_combos}] {label}  ({n_runs} runs)")

        runs: list[dict] = []
        consecutive_timeouts = 0
        TIMEOUT_CIRCUIT_BREAKER = 2

        for i in range(n_runs):
            time.sleep(delay)
            run_t0 = time.time()
            try:
                client = pool.get(model)
                if wf == "compliance":
                    result = run_compliance(client, model, temperature)
                elif wf == "dims":
                    result = run_dims(client, model, temperature)
                elif wf == "dims_ocr":
                    result = run_dims_ocr(client, model, temperature, item or "")
                elif wf == "dims_vlm":
                    result = run_dims_vlm(client, model, temperature, item or "")
                elif wf == "docs":
                    result = run_docs(client, model, temperature, item or "")
                elif wf == "docs_vlm":
                    result = run_docs_vlm(client, model, temperature, item or "")
                elif wf == "schedule_read":
                    result = run_schedule_read(client, model, temperature)
                elif wf == "schedule_write":
                    result = run_schedule_write(client, model, temperature)
                else:
                    result = {"workflow": wf, "model": model, "error": f"unknown workflow {wf}"}
            except Exception as e:
                result = {"workflow": wf, "model": model, "item": item, "error": f"call failed: {e}", "trace": traceback.format_exc()[:500]}
            run_elapsed = time.time() - run_t0

            if run_elapsed > RUN_TIMEOUT_S * 1.2 and "error" not in result:
                result.setdefault("metadata", {})["wall_clock_warning"] = (
                    f"run took {run_elapsed:.1f}s, exceeds RUN_TIMEOUT_S budget of {RUN_TIMEOUT_S}s"
                )
            runs.append(result)

            with counter_lock:
                counters["runs"] += 1
                progress = f"[{counters['runs']}/{total_runs_planned}]"

            if "error" in result:
                err_str = (result.get('error', '') or '').lower()
                is_timeout = "timed out" in err_str or "timeout" in err_str
                if is_timeout:
                    consecutive_timeouts += 1
                else:
                    consecutive_timeouts = 0
                with print_lock:
                    print(f"  {label}  {progress} run {i+1:2d}/{n_runs}: ERROR — {result.get('error', '')[:120]}")
            else:
                consecutive_timeouts = 0
                s = result["score"]
                with print_lock:
                    print(f"  {label}  {progress} run {i+1:2d}/{n_runs}: P={s['precision']:.3f}  R={s['recall']:.3f}  F1={s['f1']:.3f}")

            remaining = n_runs - (i + 1)
            if consecutive_timeouts >= TIMEOUT_CIRCUIT_BREAKER and remaining > 0:
                with print_lock:
                    print(f"  {label}  ⚠ Circuit breaker: {consecutive_timeouts} consecutive timeouts — skipping remaining {remaining} run(s).")
                for j in range(remaining):
                    skip_result = {
                        "workflow": wf, "model": model, "item": item,
                        "error": f"skipped: circuit-breaker tripped after {consecutive_timeouts} consecutive timeouts",
                    }
                    runs.append(skip_result)
                    with counter_lock:
                        counters["runs"] += 1
                        progress2 = f"[{counters['runs']}/{total_runs_planned}]"
                    with print_lock:
                        print(f"  {label}  {progress2} run {i+2+j:2d}/{n_runs}: SKIPPED (circuit breaker)")
                break

        agg = aggregate(runs)
        combo_record = {
            "workflow": wf,
            "model": model,
            "drawing": item if wf in ("dims_ocr", "dims_vlm") else None,
            "pdf":     item if wf in ("docs", "docs_vlm")  else None,
            "aggregate": agg,
            "runs": runs,
        }
        if not no_save:
            with file_lock:
                with open(RESULTS / f"run_{timestamp}.partial.jsonl", "a", encoding="utf-8") as _ckpt:
                    _ckpt.write(json.dumps(combo_record, default=str) + "\n")
        with print_lock:
            print(f"  {label}  → aggregate: P={fmt_pm(agg['precision'])}  R={fmt_pm(agg['recall'])}  F1={fmt_pm(agg['f1'])}  lat={fmt_pm(agg['latency_s'], places=1)}s")
            print()
        return combo_record

    for wf in workflows:
        wf_block: dict = {
            "name": workflow_meta[wf]["name"],
            "description": workflow_meta[wf]["description"],
            "combos": [],
        }
        report["workflows"][wf] = wf_block

        wf_models = models_for(wf)
        if wf in ("dims_ocr", "dims_vlm"):
            combos = [(model, item) for model in wf_models for item in args.drawings]
        elif wf in ("docs", "docs_vlm"):
            combos = [(model, item) for model in wf_models for item in args.pdfs]
        else:
            combos = [(model, None) for model in wf_models]

        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = {}
            for model, item in combos:
                future = executor.submit(
                    _run_combo, wf, model, item,
                    args.temperature, args.runs, args.delay,
                    pool, timestamp, args.no_save,
                    total_combos, total_runs_planned,
                    counters, counter_lock, file_lock, print_lock,
                )
                futures[future] = (model, item)

            for future in as_completed(futures):
                combo_record = future.result()
                wf_block["combos"].append(combo_record)

        if len(wf_block["combos"]) >= 2:
            print(f"--- {wf} ranked (F1 mean, descending) ---")
            ranked = sorted(
                wf_block["combos"],
                key=lambda c: (c["aggregate"]["f1"]["mean"] if c["aggregate"]["f1"]["mean"] is not None else -1),
                reverse=True,
            )
            for rank_i, c in enumerate(ranked, 1):
                ag = c["aggregate"]
                tag = c["model"] + (f" / {c.get('drawing') or c.get('pdf')}" if (c.get('drawing') or c.get('pdf')) else "")
                f1m = ag["f1"]["mean"] if ag["f1"]["mean"] is not None else 0
                print(f"  #{rank_i:>2} {tag:60s}  F1={f1m:.3f}  P={fmt_pm(ag['precision'])}  R={fmt_pm(ag['recall'])}  lat={(ag['latency_s']['mean'] or 0):.1f}s")
            print()

    print_stdout_summary(report)

    if not args.no_save:
        full_path = RESULTS / f"run_{timestamp}.json"
        full_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
        md_path = RESULTS / f"summary_{timestamp}.md"
        md_path.write_text(render_summary_md(report), encoding="utf-8")
        print(f"Saved full report: {full_path}")
        print(f"Saved markdown summary: {md_path}")
        partial = RESULTS / f"run_{timestamp}.partial.jsonl"
        if partial.is_file():
            print(f"Partial checkpoint: {partial}  (delete if not needed)")


if __name__ == "__main__":
    main()
